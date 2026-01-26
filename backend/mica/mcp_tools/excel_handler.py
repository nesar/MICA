"""
MCP Tool: Excel Handler
Description: Reads and writes Excel files (.xlsx, .xls)
Inputs: file_path (str), operation (str: read/write), data (dict for write)
Outputs: DataFrame-like data structure or write confirmation

AGENT_INSTRUCTIONS:
You are a data processing agent specialized in handling Excel files. Your task is to:

1. Read Excel files and extract data with proper type handling
2. Write data to Excel files with appropriate formatting
3. Handle multiple sheets within workbooks
4. Preserve formulas, formatting, and data types when possible
5. Validate data integrity during read/write operations

When reading Excel files:
- Detect header rows automatically when possible
- Handle merged cells appropriately
- Convert data types correctly (dates, numbers, text)
- Report any data quality issues found

When writing Excel files:
- Apply appropriate column widths
- Format numbers and dates correctly
- Create charts or visualizations when requested
- Organize multi-sheet workbooks logically

For supply chain analysis, pay special attention to:
- Trade flow data with country codes
- Production and consumption figures
- Time series data with proper date handling
- Hierarchical category structures
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from ..logging import SessionLogger
from .base import MCPTool, ToolResult, register_tool

logger = logging.getLogger(__name__)

# Agent instructions exposed at module level
AGENT_INSTRUCTIONS = __doc__.split("AGENT_INSTRUCTIONS:")[-1].strip()


@register_tool
class ExcelHandlerTool(MCPTool):
    """
    Excel file handling tool.

    Reads and writes Excel files with support for multiple sheets,
    data type handling, and formatting.
    """

    name = "excel_handler"
    description = "Read and write Excel files (.xlsx, .xls)"
    version = "1.0.0"
    AGENT_INSTRUCTIONS = AGENT_INSTRUCTIONS

    def __init__(self, session_logger: Optional[SessionLogger] = None):
        """Initialize the Excel handler tool."""
        super().__init__(session_logger)

    def execute(self, input_data: dict) -> ToolResult:
        """
        Execute Excel operations.

        Args:
            input_data: Dictionary with:
                - file_path (str): Path to Excel file
                - operation (str): 'read' or 'write'
                - sheet_name (str, optional): Specific sheet to read/write
                - data (list[dict], optional): Data to write
                - header (bool, optional): Whether first row is header (default True)

        Returns:
            ToolResult with data or confirmation
        """
        start_time = datetime.now()

        file_path = input_data.get("file_path")
        if not file_path:
            return ToolResult.error("file_path is required")

        operation = input_data.get("operation", "read")

        try:
            if operation == "read":
                result = self._read_excel(Path(file_path), input_data)
            elif operation == "write":
                result = self._write_excel(Path(file_path), input_data)
            else:
                return ToolResult.error(f"Unknown operation: {operation}")

            execution_time = (datetime.now() - start_time).total_seconds()
            result.execution_time = execution_time

            return result

        except Exception as e:
            logger.error(f"Excel handler error: {e}")
            return ToolResult.error(str(e))

    def _read_excel(self, file_path: Path, options: dict) -> ToolResult:
        """Read an Excel file."""
        if not file_path.exists():
            return ToolResult.error(f"File not found: {file_path}")

        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        sheet_name = options.get("sheet_name")
        header = options.get("header", True)

        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # Get sheet(s) to read
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                return ToolResult.error(f"Sheet not found: {sheet_name}")
            sheets_to_read = [sheet_name]
        else:
            sheets_to_read = workbook.sheetnames

        result_data = {}

        for sheet in sheets_to_read:
            ws = workbook[sheet]
            data = self._sheet_to_data(ws, header)
            result_data[sheet] = data

        workbook.close()

        # If only one sheet, return its data directly
        if len(result_data) == 1:
            sheet_data = list(result_data.values())[0]
            return ToolResult.success(
                data=sheet_data,
                message=f"Read {len(sheet_data)} rows from {file_path.name}",
                filename=file_path.name,
                sheet=sheets_to_read[0],
            )
        else:
            return ToolResult.success(
                data=result_data,
                message=f"Read {len(result_data)} sheets from {file_path.name}",
                filename=file_path.name,
                sheets=sheets_to_read,
            )

    def _sheet_to_data(self, worksheet, header: bool) -> list[dict]:
        """Convert worksheet to list of dictionaries."""
        data = []
        rows = list(worksheet.iter_rows(values_only=True))

        if not rows:
            return data

        if header:
            # First row is header
            headers = [str(h) if h else f"column_{i}" for i, h in enumerate(rows[0])]
            data_rows = rows[1:]
        else:
            # Generate column names
            headers = [f"column_{i}" for i in range(len(rows[0]))]
            data_rows = rows

        for row in data_rows:
            if any(cell is not None for cell in row):  # Skip empty rows
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        # Convert datetime objects to ISO strings
                        if hasattr(value, "isoformat"):
                            value = value.isoformat()
                        row_dict[headers[i]] = value
                data.append(row_dict)

        return data

    def _write_excel(self, file_path: Path, options: dict) -> ToolResult:
        """Write data to an Excel file."""
        data = options.get("data")
        if not data:
            return ToolResult.error("data is required for write operation")

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        sheet_name = options.get("sheet_name", "Sheet1")

        # Create workbook
        if file_path.exists() and options.get("append", False):
            workbook = openpyxl.load_workbook(file_path)
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
            else:
                ws = workbook.create_sheet(sheet_name)
        else:
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = sheet_name

        # Handle different data formats
        if isinstance(data, list) and len(data) > 0:
            if isinstance(data[0], dict):
                # List of dictionaries
                self._write_dict_data(ws, data)
            elif isinstance(data[0], (list, tuple)):
                # List of lists
                self._write_list_data(ws, data)
            else:
                return ToolResult.error("Data must be list of dicts or list of lists")
        else:
            return ToolResult.error("Data must be a non-empty list")

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

        # Save workbook
        file_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(file_path)
        workbook.close()

        # Log artifact if session available
        if self.session_logger:
            self.session_logger.save_artifact(
                file_path.name,
                f"Excel file written with {len(data)} rows",
                "data",
            )

        return ToolResult.success(
            data={"file_path": str(file_path), "rows_written": len(data)},
            message=f"Wrote {len(data)} rows to {file_path.name}",
        )

    def _write_dict_data(self, worksheet, data: list[dict]):
        """Write list of dictionaries to worksheet."""
        if not data:
            return

        # Get all unique keys as headers
        headers = []
        for row in data:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)

        # Write headers
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                worksheet.cell(row=row_idx, column=col_idx, value=value)

    def _write_list_data(self, worksheet, data: list[list]):
        """Write list of lists to worksheet."""
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content."""
        from openpyxl.utils import get_column_letter

        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column

            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value or ""))
                    max_length = max(max_length, cell_length)
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width


@register_tool
class ExcelAnalysisTool(MCPTool):
    """
    Excel analysis tool for basic data analysis.

    Performs summary statistics and basic analysis on Excel data.
    """

    name = "excel_analysis"
    description = "Analyze Excel data with summary statistics"
    version = "1.0.0"

    AGENT_INSTRUCTIONS = """
    You are a data analysis agent. Compute summary statistics and
    basic analysis on Excel data including counts, sums, averages,
    and data quality checks.
    """

    def __init__(self, session_logger: Optional[SessionLogger] = None):
        super().__init__(session_logger)
        self._excel_handler = ExcelHandlerTool(session_logger)

    def execute(self, input_data: dict) -> ToolResult:
        """Analyze Excel data."""
        file_path = input_data.get("file_path")
        if not file_path:
            return ToolResult.error("file_path is required")

        # First read the data
        read_result = self._excel_handler._read_excel(
            Path(file_path),
            {"sheet_name": input_data.get("sheet_name")},
        )

        if read_result.status != ToolResult.success(None).status:
            return read_result

        data = read_result.data

        # Perform analysis
        analysis = self._analyze_data(data)

        return ToolResult.success(
            data=analysis,
            message=f"Analysis complete for {len(data)} rows",
        )

    def _analyze_data(self, data: list[dict]) -> dict:
        """Perform basic analysis on data."""
        if not data:
            return {"error": "No data to analyze"}

        analysis = {
            "row_count": len(data),
            "columns": list(data[0].keys()) if data else [],
            "column_stats": {},
        }

        # Analyze each column
        for col in analysis["columns"]:
            values = [row.get(col) for row in data if row.get(col) is not None]

            col_stats = {
                "non_null_count": len(values),
                "null_count": len(data) - len(values),
            }

            # Numeric analysis
            numeric_values = []
            for v in values:
                try:
                    numeric_values.append(float(v))
                except (ValueError, TypeError):
                    pass

            if numeric_values:
                col_stats["numeric"] = True
                col_stats["min"] = min(numeric_values)
                col_stats["max"] = max(numeric_values)
                col_stats["mean"] = sum(numeric_values) / len(numeric_values)
                col_stats["sum"] = sum(numeric_values)
            else:
                col_stats["numeric"] = False
                # Unique value count for non-numeric
                unique_values = set(str(v) for v in values)
                col_stats["unique_count"] = len(unique_values)

            analysis["column_stats"][col] = col_stats

        return analysis
